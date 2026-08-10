#!/usr/bin/env python3
"""
Step 1 - Excel Question Bank Parser (灵活解析版)
自动检测 Excel 题库结构，不限固定 Sheet 名/列位/格式

输入: 任意格式的 360 环评题库 Excel
输出: question_bank.json (标准化结构)

支持的变体:
  - Sheet 名不固定（自动识别"评价表""评价人""辅助"等）
  - 列位不固定（自动检测表头行和数据列）
  - 题目编号支持: 一、二、三 / 1. 2. 3. / ① ② ③
  - 维度名: 任意中文短词(2~8字)
  - 评分选项: N分格式 + 无法判断

Usage:
  python parse_excel.py <excel_path> [--output <json_path>]
"""

import json, re, sys, os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl


# ============================================================
# Sheet 自动识别
# ============================================================

def _identify_sheets(wb) -> dict:
    """识别三个 Sheet 的角色: question_bank, evaluators, scoring"""
    result = {'question_bank': None, 'evaluators': None, 'scoring': None}
    remaining = list(wb.sheetnames)

    # 优先级1: 按名称关键词匹配
    for name in remaining[:]:
        if not result['question_bank']:
            if any(kw in name for kw in ['评价表', '胜任力', '问卷', '360', '环评', 'questions']):
                result['question_bank'] = name
                remaining.remove(name)
                continue
        if not result['evaluators']:
            if any(kw in name for kw in ['评价人', '名单', '人员', 'rater', 'evaluator']):
                result['evaluators'] = name
                remaining.remove(name)
                continue
        if not result['scoring']:
            if any(kw in name for kw in ['辅助', '勿删', '选项', '评分', 'scoring', 'helper']):
                result['scoring'] = name
                remaining.remove(name)
                continue

    # 优先级2: 按内容特征识别
    for name in remaining[:]:
        ws = wb[name]
        sample = _sample_rows(ws, 10)

        if not result['question_bank'] and _looks_like_question_bank(sample):
            result['question_bank'] = name
            remaining.remove(name)
            continue
        if not result['evaluators'] and _looks_like_evaluators(sample, ws):
            result['evaluators'] = name
            remaining.remove(name)
            continue
        if not result['scoring'] and _looks_like_scoring(sample):
            result['scoring'] = name
            remaining.remove(name)
            continue

    # 优先级3: 兜底——第一个未识别的 Sheet 做题库
    if not result['question_bank'] and remaining:
        result['question_bank'] = remaining[0]

    return result


def _sample_rows(ws, n=10):
    return [str(ws.cell(row=r, column=1).value or '') for r in range(1, min(ws.max_row + 1, n + 1))]


def _looks_like_question_bank(sample: list) -> bool:
    """检测是否像题库 Sheet: 有维度名(短中文)或编号题目"""
    dim_like = sum(1 for t in sample if re.match(r'^[\u4e00-\u9fa5]{2,8}$', t.strip()))
    q_like = sum(1 for t in sample if _detect_question(t))
    return (dim_like >= 2) or (q_like >= 3) or (dim_like + q_like >= 3)


def _looks_like_evaluators(sample: list, ws) -> bool:
    """检测是否像评价人清单: 有级别/关系列 + 姓名列"""
    # 看第一行是否有多个列
    headers = [str(ws.cell(row=1, column=c).value or '') for c in range(1, min(ws.max_column + 1, 5))]
    joined = '|'.join(headers)
    if any(kw in joined for kw in ['级别', '关系', '人员', '姓名', '职位', 'level', 'name']):
        return True
    # 看数据行是否有 上级/平级/下级
    for r in range(2, min(ws.max_row + 1, 6)):
        first_col = str(ws.cell(row=r, column=1).value or '')
        if any(kw in first_col for kw in ['上级', '平级', '下级']):
            return True
    return False


def _looks_like_scoring(sample: list) -> bool:
    """检测是否像评分选项 Sheet: 有 N分 或 无法判断"""
    score_count = sum(1 for t in sample if re.match(r'\d+分', t.strip()))
    na_count = sum(1 for t in sample if '无法判断' in t or 'N/A' in t)
    return score_count >= 2 or na_count >= 1


# ============================================================
# 题库 Sheet 解析
# ============================================================

def _parse_question_bank_sheet(ws) -> tuple:
    """解析题库 Sheet，返回 (meta, dimensions)"""
    rows = []
    for r in range(1, ws.max_row + 1):
        val_a = ws.cell(row=r, column=1).value
        val_b = ws.cell(row=r, column=2).value
        rows.append((str(val_a).strip() if val_a else '', str(val_b).strip() if val_b else ''))

    # --- 元数据检测 ---
    meta = {'subject': '', 'title': '', 'table_title': ''}
    meta_start = 0
    table_title_found = False
    for i, (a, b) in enumerate(rows):
        subj = _extract_after(a, ['被评价人：', '被评价人:', '评价对象：', '姓名：', 'Subject:'])
        ttl = _extract_after(a, ['职级：', '职级:', '职位：', 'Title:'])
        if subj:
            meta['subject'] = subj
            meta_start = max(meta_start, i + 1)
            continue
        if ttl:
            meta['title'] = ttl
            meta_start = max(meta_start, i + 1)
            continue
        # 标题行: 在遇到维度/题目之前，首个长度>=4的非meta非问号行
        if not table_title_found and not _detect_question(a) and len(a) >= 4:
            # 确认不是维度名（短中文）
            if not re.match(r'^[\u4e00-\u9fa5·]{2,6}$', a):
                meta['table_title'] = a
                table_title_found = True
                meta_start = max(meta_start, i + 1)
                continue
        # 遇到维度或题目 → 停止元数据检测
        if re.match(r'^[\u4e00-\u9fa5·]{2,8}$', a) or _detect_question(a):
            meta_start = i
            break

    # --- 维度和题目 ---
    dimensions = []
    current_dim = None
    q_idx = 0
    # 用于跨多单元格查找题目标签（有些题库把标签放 B 列）
    pending_tag = ''

    for i in range(meta_start, len(rows)):
        a, b = rows[i]

        # 跳过空行
        if not a:
            pending_tag = ''
            continue

        # 维度检测: 纯中文短词(2-8字)，无数号/编号
        if re.match(r'^[\u4e00-\u9fa5·]{2,8}$', a) and not _detect_question(a):
            # 排除 meta 行
            if any(kw in a for kw in ['被评价人', '职级', '姓名', '评价人']):
                continue
            current_dim = {'name': a, 'questions': []}
            dimensions.append(current_dim)
            pending_tag = ''
            continue

        # 题目检测
        q_info = _parse_question_line(a)
        if q_info:
            q_idx += 1
            tag = q_info['tag']
            # 如果 A 列没标签，尝试从 B 列或下行获取
            if not tag and b:
                tag = _extract_tag_from_text(b)
            question = {
                'id': f'Q{q_idx}',
                'order_cn': q_info.get('order_cn', str(q_idx)),
                'description': q_info['description'],
                'tag': tag or f'题{q_idx}'
            }
            if current_dim is not None:
                current_dim['questions'].append(question)
            else:
                # 没有维度时，自动生成"未分组"维度
                if not dimensions or dimensions[-1]['name'] != '综合评价':
                    current_dim = {'name': '综合评价', 'questions': []}
                    dimensions.append(current_dim)
                current_dim['questions'].append(question)
            pending_tag = ''
        elif pending_tag and current_dim and current_dim['questions']:
            # 上一行题目没有标签，这一行是标签补充
            current_dim['questions'][-1]['tag'] = pending_tag
            pending_tag = ''

    # 清理 B 列残留标签
    _merge_column_b_tags(ws, dimensions)

    return meta, dimensions


def _merge_column_b_tags(ws, dimensions):
    """部分题库把题目标签放在 B 列，合并到对应题目的 tag 中"""
    # 简化处理: 遍历所有维度题目，如果 tag 等于题号，尝试从 B 列找替换
    pass  # 当前题库模板无此情况，预留扩展


def _extract_after(text: str, prefixes: list) -> str:
    """从文本中提取前缀后的内容"""
    for p in prefixes:
        if p in text:
            return text.split(p, 1)[1].strip()
    return ''


def _extract_tag_from_text(text: str) -> str:
    """从文本中提取可能的标签（括号内的短词）"""
    m = re.search(r'[（(]([\u4e00-\u9fa5\w·]{2,12})[）)]', text)
    if m:
        return m.group(1)
    # 如果直接是短词
    if re.match(r'^[\u4e00-\u9fa5·]{2,6}$', text.strip()):
        return text.strip()
    return ''


# ============================================================
# 题目行解析
# ============================================================

# 中文序号 → 数字
_CN_NUM = {'一':1,'二':2,'三':3,'四':4,'五':5,'六':6,'七':7,'八':8,'九':9,'十':10,
           '十一':11,'十二':12,'十三':13,'十四':14,'十五':15,'十六':16,'十七':17,'十八':18,'十九':19,'二十':20}
# 编号序号
_CIRCLE_NUM = {f'①':1,f'②':2,f'③':3,f'④':4,f'⑤':5,f'⑥':6,f'⑦':7,f'⑧':8,f'⑨':9,f'⑩':10}

_QUESTION_PATTERNS = [
    # 格式: "一、描述。（标签）" 或 "一、描述（标签）" 或 "一、描述"
    re.compile(r'^([一二三四五六七八九十]+)、(.+?)。?[（(](.+?)[）)]$'),
    re.compile(r'^([一二三四五六七八九十]+)、(.+?)[（(](.+?)[）)]$'),
    re.compile(r'^([一二三四五六七八九十]+)、(.+?)$'),
    # 格式: "1. 描述" 或 "1、描述"
    re.compile(r'^(\d+)[.、]\s*(.+?)。?[（(](.+?)[）)]\s*$'),
    re.compile(r'^(\d+)[.、]\s*(.+?)\s*$'),
    # 格式: "①描述"
    re.compile(r'^([①②③④⑤⑥⑦⑧⑨⑩])\s*(.+?)。?[（(](.+?)[）)]\s*$'),
    re.compile(r'^([①②③④⑤⑥⑦⑧⑨⑩])\s*(.+?)\s*$'),
]


def _detect_question(text: str) -> bool:
    """判断文本是否像一道题目"""
    for pat in _QUESTION_PATTERNS:
        if pat.match(text):
            return True
    return False


def _parse_question_line(text: str) -> dict | None:
    """解析题目行，返回 {order_cn, description, tag} 或 None"""
    for pat in _QUESTION_PATTERNS:
        m = pat.match(text)
        if m:
            order_str = m.group(1)
            desc = m.group(2).strip()
            # 去掉描述末尾的句号残留
            desc = re.sub(r'[。；;]+$', '', desc).strip()
            tag = _normalize_tag(m.group(3).strip()) if m.lastindex and m.lastindex >= 3 else ''
            return {
                'order_cn': order_str,
                'description': desc,
                'tag': tag
            }
    return None


def _normalize_tag(tag: str) -> str:
    """标准化标签: 去掉多余括弧和空白"""
    tag = tag.strip()
    tag = re.sub(r'^[（(]|[）)]$', '', tag).strip()
    return tag


# ============================================================
# 评价人 Sheet 解析
# ============================================================

def _parse_evaluator_sheet(ws) -> list:
    """解析评价人 Sheet，返回 [{level, name, position}]"""
    evaluators = []

    # 先检测表头 → 确定列映射
    col_map = {}  # 级别/人员/职位 → 列号(1-based)
    for c in range(1, min(ws.max_column + 1, 5)):
        header = str(ws.cell(row=1, column=c).value or '')
        if any(kw in header for kw in ['级别', '关系', 'level', '上级', '平级']):
            col_map['level'] = c
        elif any(kw in header for kw in ['人员', '姓名', '名字', 'name']):
            col_map['person'] = c
        elif any(kw in header for kw in ['职位', '岗位', 'title', 'position']):
            col_map['position'] = c

    # 如果表头没检测到，用默认映射
    if not col_map:
        col_map = {'level': 1, 'person': 2, 'position': 3}

    level_col = col_map.get('level', 1)
    person_col = col_map.get('person', 2)
    pos_col = col_map.get('position', 3)

    current_level = None
    for r in range(2, ws.max_row + 1):
        level = str(ws.cell(row=r, column=level_col).value or '').strip()
        person = str(ws.cell(row=r, column=person_col).value or '').strip()
        position = str(ws.cell(row=r, column=pos_col).value or '').strip()

        # 跳过纯表头行
        if person in ['人员', '姓名', '']:
            if level and level not in ['级别', '']:
                current_level = level
            continue

        if level and level not in ['级别']:
            current_level = level

        evaluators.append({
            'level': current_level or '未指定',
            'name': person,
            'position': position
        })

    return evaluators


# ============================================================
# 评分选项 Sheet 解析
# ============================================================

def _parse_scoring_sheet(ws) -> list:
    """解析评分选项 Sheet"""
    options = []
    for r in range(1, ws.max_row + 1):
        # 遍历所有列找评分描述
        for c in range(1, ws.max_column + 1):
            t = str(ws.cell(row=r, column=c).value or '').strip()
            if not t:
                continue
            m = re.match(r'^(\d+)分', t)
            if m:
                options.append({'score': int(m.group(1)), 'text': t})
                break
            elif '无法判断' in t or 'N/A' in t or '不适用' in t:
                options.append({'score': None, 'text': t})
                break

    # 按分数排序
    options.sort(key=lambda x: (x['score'] is not None, -((x['score'] or 0))), reverse=True)
    return options


# ============================================================
# 主入口
# ============================================================

def parse_excel(excel_path: str) -> dict:
    """解析题库 Excel，返回标准化 JSON"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_map = _identify_sheets(wb)

    # 解析题库
    qb_name = sheet_map['question_bank']
    if not qb_name:
        raise ValueError(f'未找到题库 Sheet，文件包含: {wb.sheetnames}')
    ws_qb = wb[qb_name]
    meta, dimensions = _parse_question_bank_sheet(ws_qb)

    # 解析评价人
    evaluators = []
    ev_name = sheet_map['evaluators']
    if ev_name:
        evaluators = _parse_evaluator_sheet(wb[ev_name])

    # 解析评分选项
    scoring_options = []
    sc_name = sheet_map['scoring']
    if sc_name:
        scoring_options = _parse_scoring_sheet(wb[sc_name])

    # 如果没解析到评分选项，用默认值
    if not scoring_options:
        scoring_options = [
            {'score': 5, 'text': '5分：持续高频展现此能力，行动结果卓越。'},
            {'score': 4, 'text': '4分：经常展现此能力，行动结果优良。'},
            {'score': 3, 'text': '3分：能够展现此能力，行动结果合格。'},
            {'score': 2, 'text': '2分：偶尔展现此能力，行动结果有待提高。'},
            {'score': 1, 'text': '1分：从未展现此能力，无法满足工作要求。'},
            {'score': None, 'text': '无法判断'}
        ]

    q_idx = sum(len(d['questions']) for d in dimensions)
    return {
        'meta': meta,
        'dimensions': dimensions,
        'total_questions': q_idx,
        'evaluators': evaluators,
        'scoring_options': scoring_options,
        'source_file': str(Path(excel_path).name),
        '_sheets_used': {k: v for k, v in sheet_map.items() if v}
    }


def main():
    if len(sys.argv) < 2:
        print("Usage: python parse_excel.py <excel_path> [--output <json_path>]")
        sys.exit(1)

    excel_path = sys.argv[1]
    output_path = None
    if '--output' in sys.argv:
        idx = sys.argv.index('--output')
        output_path = sys.argv[idx + 1] if idx + 1 < len(sys.argv) else None

    result = parse_excel(excel_path)

    if output_path:
        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f'✅ 题库结构已保存到: {output_path}')

    sheets = result.pop('_sheets_used', {})
    print(f"\n📋 题库: {result['meta'].get('table_title', '未知')}")
    print(f"   被评价人: {result['meta'].get('subject', '未知')} ({result['meta'].get('title', '')})")
    print(f"   识别到的 Sheet: 题库={sheets.get('question_bank','?')}  评价人={sheets.get('evaluators','无')}  评分={sheets.get('scoring','默认')}")
    print(f"   维度数: {len(result['dimensions'])} | 题目数: {result['total_questions']}")
    print(f"   评价人数: {len(result['evaluators'])}")
    print(f"\n--- 维度分布 ---")
    for d in result['dimensions']:
        print(f"   {d['name']}: {len(d['questions'])} 题")
    if result['evaluators']:
        print(f"\n--- 评价人 ---")
        for e in result['evaluators']:
            print(f"   [{e['level']}] {e['name']} ({e['position']})")


if __name__ == '__main__':
    main()
